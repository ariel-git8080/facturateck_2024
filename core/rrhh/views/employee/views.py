import json
from datetime import datetime
from io import BytesIO

import xlsxwriter
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Group
from django.db import transaction
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView, TemplateView, View
from openpyxl import load_workbook

from config import settings
from core.rrhh.forms import EmployeeForm, User, Employee, EmployeeUserForm
from core.security.mixins import GroupModuleMixin, GroupPermissionMixin


class EmployeeListView(GroupPermissionMixin, TemplateView):
    template_name = 'employee/list.html'
    permission_required = 'view_employee'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search':
                data = []
                for i in Employee.objects.filter():
                    data.append(i.toJSON())
            elif action == 'upload_excel':
                with transaction.atomic():
                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    excel = workbook[workbook.sheetnames[0]]
                    for row in range(2, excel.max_row + 1):
                        code = excel.cell(row=row, column=1).value
                        employee = Employee.objects.filter(code=code).first()
                        if employee is None:
                            employee = Employee()
                        if employee and hasattr(employee, 'user'):
                            user = employee.user
                        else:
                            user = User()
                        user.names = excel.cell(row=row, column=2).value
                        dni = excel.cell(row=row, column=3).value
                        user.is_active = int(excel.cell(row=row, column=8).value)
                        user.username = dni
                        user.create_or_update_password(user.username)
                        user.save()
                        employee.code = code
                        employee.user_id = user.id
                        employee.dni = dni
                        employee.position = employee.get_or_create_position(name=excel.cell(row=row, column=4).value)
                        employee.area = employee.get_or_create_area(name=excel.cell(row=row, column=6).value)
                        employee.hiring_date = excel.cell(row=row, column=5).value
                        employee.remuneration = str(excel.cell(row=row, column=7).value).replace('.', '')
                        employee.save()
                        group = Group.objects.get(pk=settings.GROUPS['employee'])
                        if not user.groups.filter(id=group.id).exists():
                            user.groups.add(group)
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Empleados'
        context['create_url'] = reverse_lazy('employee_create')
        return context


class EmployeeCreateView(GroupPermissionMixin, CreateView):
    model = Employee
    template_name = 'employee/create.html'
    form_class = EmployeeForm
    success_url = reverse_lazy('employee_list')
    permission_required = 'add_employee'

    def get_form_user(self):
        form = EmployeeUserForm()
        if self.request.POST or self.request.FILES:
            form = EmployeeUserForm(self.request.POST, self.request.FILES)
        return form

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'add':
                with transaction.atomic():
                    form1 = self.get_form_user()
                    form2 = self.get_form()
                    if form1.is_valid() and form2.is_valid():
                        user = form1.save(commit=False)
                        user.username = form2.cleaned_data['dni']
                        user.set_password(user.username)
                        user.save()
                        user.groups.add(Group.objects.get(pk=settings.GROUPS['employee']))
                        form_employee = form2.save(commit=False)
                        form_employee.user = user
                        form_employee.save()
                    else:
                        if not form1.is_valid():
                            data['error'] = form1.errors
                        elif not form2.is_valid():
                            data['error'] = form2.errors
            elif action == 'validate_data':
                data = {'valid': True}
                queryset = Employee.objects.all()
                pattern = request.POST['pattern']
                parameter = request.POST['parameter'].strip()
                if pattern == 'dni':
                    data['valid'] = not queryset.filter(dni=parameter).exists()
                elif pattern == 'code':
                    data['valid'] = not queryset.filter(code=parameter).exists()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['title'] = 'Nuevo registro de un Empleado'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        context['frmUser'] = self.get_form_user()
        return context


class EmployeeUpdateView(GroupPermissionMixin, UpdateView):
    model = Employee
    template_name = 'employee/create.html'
    form_class = EmployeeForm
    success_url = reverse_lazy('employee_list')
    permission_required = 'change_employee'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def get_form_user(self):
        form = EmployeeUserForm(instance=self.request.user)
        if self.request.POST or self.request.FILES:
            form = EmployeeUserForm(self.request.POST, self.request.FILES, instance=self.object.user)
        return form

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'edit':
                with transaction.atomic():
                    form1 = self.get_form_user()
                    form2 = self.get_form()
                    if form1.is_valid() and form2.is_valid():
                        user = form1.save(commit=False)
                        user.save()
                        form_employee = form2.save(commit=False)
                        form_employee.user = user
                        form_employee.save()
                    else:
                        if not form1.is_valid():
                            data['error'] = form1.errors
                        elif not form2.is_valid():
                            data['error'] = form2.errors
            elif action == 'validate_data':
                data = {'valid': True}
                queryset = Employee.objects.all().exclude(id=self.object.employee.id)
                pattern = request.POST['pattern']
                parameter = request.POST['parameter'].strip()
                if pattern == 'dni':
                    data['valid'] = not queryset.filter(dni=parameter).exists()
                elif pattern == 'code':
                    data['valid'] = not queryset.filter(code=parameter).exists()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['title'] = 'Edición de un Empleado'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        context['frmUser'] = EmployeeUserForm(instance=self.object.user)
        return context


class EmployeeDeleteView(GroupPermissionMixin, DeleteView):
    model = Employee
    template_name = 'delete.html'
    success_url = reverse_lazy('employee_list')
    permission_required = 'delete_employee'

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            self.get_object().delete()
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Notificación de eliminación'
        context['list_url'] = self.success_url
        return context


class EmployeeUpdateProfileView(GroupModuleMixin, UpdateView):
    model = Employee
    template_name = 'employee/profile.html'
    form_class = EmployeeForm
    success_url = settings.LOGIN_REDIRECT_URL

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        return self.request.user.employee

    def get_form(self, form_class=None):
        form = super(EmployeeUpdateProfileView, self).get_form(form_class)
        for name in ['dni', 'code', 'position', 'area', 'hiring_date', 'remuneration']:
            form.fields[name].widget.attrs['readonly'] = True
        return form

    def get_form_user(self):
        form = EmployeeUserForm(instance=self.request.user)
        if self.request.POST or self.request.FILES:
            form = EmployeeUserForm(self.request.POST, self.request.FILES, instance=self.request.user)
        for name in ['names']:
            form.fields[name].widget.attrs['readonly'] = True
            form.fields[name].required = False
        return form

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'edit':
                with transaction.atomic():
                    form1 = self.get_form_user()
                    form2 = self.get_form()
                    if form1.is_valid() and form2.is_valid():
                        user = form1.save(commit=False)
                        user.save()
                        form_employee = form2.save(commit=False)
                        form_employee.user = user
                        form_employee.save()
                    else:
                        if not form1.is_valid():
                            data['error'] = form1.errors
                        elif not form2.is_valid():
                            data['error'] = form2.errors
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['title'] = 'Edición de Perfil'
        context['list_url'] = self.success_url
        context['action'] = 'edit'
        context['frmUser'] = self.get_form_user()
        return context


class EmployeeExportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        try:
            headers = {
                'Código': 35,
                'Nombres': 50,
                'Número de documento': 35,
                'Cargo': 40,
                'Fecha de ingreso': 35,
                'Área': 40,
                'Remuneración': 40,
                'Estado': 20,
            }

            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('empleados')
            cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
            row_format = workbook.add_format({'align': 'center', 'border': 1})
            index = 0
            for name, width in headers.items():
                worksheet.set_column(first_col=index, last_col=index, width=width)
                worksheet.write(0, index, name, cell_format)
                index += 1
            row = 1
            for employee in Employee.objects.filter().order_by('id'):
                worksheet.write(row, 0, employee.code, row_format)
                worksheet.write(row, 1, employee.user.names, row_format)
                worksheet.write(row, 2, employee.dni, row_format)
                worksheet.write(row, 3, employee.position.name, row_format)
                worksheet.write(row, 4, employee.hiring_date_format(), row_format)
                worksheet.write(row, 5, employee.area.name, row_format)
                worksheet.write(row, 6, float(employee.remuneration), row_format)
                worksheet.write(row, 7, 1 if employee.user.is_active else 0, row_format)
                row += 1
            workbook.close()
            output.seek(0)
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f"attachment; filename=EMPLEADOS_{datetime.now().date().strftime('%d_%m_%Y')}.xlsx"
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('employee_list'))
