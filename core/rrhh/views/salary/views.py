import json
from datetime import datetime
from io import BytesIO

import xlsxwriter
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Sum, Q, FloatField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse_lazy
from django.views.generic import FormView, CreateView
from django.views.generic.base import View, TemplateView
from openpyxl import load_workbook

from core.pos.utilities import printer
from core.rrhh.forms import SalaryForm, Salary, SalaryDetail, SalaryHeadings, Employee, Headings, MONTHS
from core.security.mixins import GroupPermissionMixin


class SalaryListView(GroupPermissionMixin, FormView):
    form_class = SalaryForm
    template_name = 'salary/admin/list.html'
    permission_required = 'view_salary'

    def get_form(self, form_class=None):
        form = SalaryForm()
        form.fields['year'].initial = datetime.now().date().year
        return form

    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'search':
                data = []
                year = request.POST['year']
                month = request.POST['month']
                pks = json.loads(request.POST['pks'])
                queryset = SalaryDetail.objects.filter()
                if len(year):
                    queryset = queryset.filter(salary__year=year)
                if len(month):
                    queryset = queryset.filter(salary__month=month)
                if len(pks):
                    queryset = queryset.filter(employee_id__in=pks)
                for i in queryset:
                    data.append(i.toJSON())
            elif action == 'search_detail_headings':
                data = []
                detail = SalaryDetail.objects.get(pk=request.POST['id'])
                for i in detail.salaryheadings_set.filter(headings__type='haberes', valor__gt=0).order_by('headings__order'):
                    data.append([i.headings.name, i.get_cant(), i.get_valor_format(), '---'])
                for i in detail.salaryheadings_set.filter(headings__type='descuentos', valor__gt=0).order_by('headings__order'):
                    data.append([i.headings.name, i.get_cant(), '---', i.get_valor_format()])
                data.append(['Subtotal de Ingresos', '---', '---', detail.get_income_format()])
                data.append(['Subtotal de Descuentos', '---', '---', detail.get_expenses_format()])
                data.append(['Total a recibir', '---', '---', detail.get_total_amount_format()])
            elif action == 'upload_excel':
                with transaction.atomic():
                    year = int(request.POST['year'])
                    month = int(request.POST['month'])
                    archive = request.FILES['archive']
                    workbook = load_workbook(archive, keep_vba=True, read_only=True, data_only=True, keep_links=True)
                    wb = workbook[workbook.sheetnames[0]]
                    columns = []
                    for cell in wb.iter_rows(min_row=1, max_row=1, min_col=7):
                        for row in cell:
                            columns.append(row.value)
                    max_column = wb.max_column - 1
                    for cell in wb.iter_rows(min_row=2, max_row=wb.max_row):
                        row = cell[0].row
                        employee = Employee.objects.get(code=wb.cell(row=row, column=1).value)
                        salary_detail = SalaryDetail.objects.filter(employee=employee, salary__year=year, salary__month=month).first()
                        if salary_detail is None:
                            salary_detail = SalaryDetail()
                            salary_detail.salary = Salary.objects.get_or_create(year=year, month=month)[0]
                            salary_detail.employee = employee
                            salary_detail.save()
                        index = 0
                        position = 7
                        salary_detail.salaryheadings_set.all().delete()
                        salary_detail.salaryheadings_set.all().delete()
                        while position < max_column:
                            code = columns[index]
                            if code in 'Subtotal':
                                position += 1
                                index += 1
                                continue
                            detail = SalaryHeadings()
                            detail.salary_detail_id = salary_detail.id
                            if code.__contains__('Cantidad.'):
                                code = code.split('.')[-1]
                                detail.headings = Headings.objects.get(code=code)
                                detail.cant = wb.cell(row=row, column=position).value
                                valor = str(wb.cell(row=row, column=position + 1).value)
                                detail.valor = valor.replace('.', '')
                                index += 2
                                position += 2
                            else:
                                detail.headings = Headings.objects.get(name=code)
                                valor = str(wb.cell(row=row, column=position).value)
                                detail.valor = valor.replace('.', '')
                                index += 1
                                position += 1
                            detail.save()
                        salary_detail.income = salary_detail.salaryheadings_set.filter(headings__type='haberes').aggregate(result=Coalesce(Sum('valor'), 0.00, output_field=FloatField()))['result']
                        salary_detail.expenses = salary_detail.salaryheadings_set.filter(headings__type='descuentos').aggregate(result=Coalesce(Sum('valor'), 0.00, output_field=FloatField()))['result']
                        salary_detail.total_amount = float(salary_detail.income) - float(salary_detail.expenses)
                        salary_detail.save()
            elif action == 'search_employee':
                data = []
                term = request.POST['term']
                for i in Employee.objects.filter(Q(user__names__icontains=term) | Q(dni__icontains=term) | Q(code__icontains=term)).order_by('user__names')[0:10]:
                    item = i.toJSON()
                    item['text'] = i.get_full_name()
                    data.append(item)
            elif action == 'remove_salaries':
                request.session['salaries'] = {}
                pks = json.loads(request.POST['pks'])
                request.session['salaries']['year'] = int(request.POST['year'])
                month = request.POST['month']
                request.session['salaries']['month'] = {}
                if len(month):
                    request.session['salaries']['month']['id'] = int(month)
                    request.session['salaries']['month']['name'] = MONTHS[int(month)][1]
                request.session['salaries']['employees'] = {}
                if len(pks):
                    request.session['salaries']['employees'] = Employee.objects.filter(id__in=pks)
                data['url'] = str(reverse_lazy('salary_delete'))
            elif action == 'export_salaries_pdf':
                year = request.POST['year']
                month = request.POST['month']
                pks = json.loads(request.POST['pks'])
                queryset = SalaryDetail.objects.filter(salary__year=year)
                if len(month):
                    queryset = queryset.filter(salary__month=month)
                if len(pks):
                    queryset = queryset.filter(employee_id__in=pks)
                context = {
                    'salaries': queryset,
                    'company': request.tenant.company,
                    'prints': [1, 2],
                    'date_joined': datetime.now().date()
                }
                pdf_file = printer.create_pdf(context=context, template_name='salary/format/format2.html')
                return HttpResponse(pdf_file, content_type='application/pdf')
            elif action == 'export_salaries_excel':
                year = request.POST['year']
                month = request.POST['month']
                pks = json.loads(request.POST['pks'])
                queryset = SalaryDetail.objects.filter(salary__year=year)
                if len(month):
                    queryset = queryset.filter(salary__month=month)
                if len(pks):
                    queryset = queryset.filter(employee_id__in=pks)
                headers = {
                    'Código': 15,
                    'Empleado': 35,
                    'Sección': 35,
                    'Cargo': 35,
                    'Número de documento': 35,
                    'Fecha de ingreso': 35,
                }
                headings = Headings.objects.filter()
                for i in headings.filter(type='haberes').order_by('order'):
                    if i.has_quantity:
                        key = f'Cantidad {i.name}'
                        headers[key] = 45
                    headers[i.name] = 55
                headers['Subtotal'] = 50
                for i in headings.filter(type='descuentos').order_by('order'):
                    if i.has_quantity:
                        key = f'Cantidad {i.name}'
                        headers[key] = 45
                    headers[i.name] = 55
                headers['Total Descuento'] = 50
                headers['Total a Cobrar'] = 40
                output = BytesIO()
                workbook = xlsxwriter.Workbook(output)
                worksheet = workbook.add_worksheet('planilla')
                cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
                row_format = workbook.add_format({'align': 'center', 'border': 1})
                index = 0
                for name, width in headers.items():
                    worksheet.set_column(first_col=0, last_col=index, width=width)
                    worksheet.write(0, index, name, cell_format)
                    index += 1
                row = 1
                for salary_detail in queryset.order_by('employee'):
                    worksheet.write(row, 0, salary_detail.employee.code, row_format)
                    worksheet.write(row, 1, salary_detail.employee.user.names, row_format)
                    worksheet.write(row, 2, salary_detail.employee.area.name, row_format)
                    worksheet.write(row, 3, salary_detail.employee.position.name, row_format)
                    worksheet.write(row, 4, salary_detail.employee.dni, row_format)
                    worksheet.write(row, 5, salary_detail.employee.hiring_date_format(), row_format)
                    index = 5
                    for heading in headings.filter(type='haberes').order_by('order'):
                        salary_headings = salary_detail.salaryheadings_set.filter(headings_id=heading.id).first()
                        if salary_headings:
                            if heading.has_quantity:
                                worksheet.write(row, index + 1, salary_headings.get_cant(), row_format)
                                worksheet.write(row, index + 2, salary_headings.get_valor_format(), row_format)
                                index += 2
                            else:
                                worksheet.write(row, index + 1, salary_headings.get_valor_format(), row_format)
                                index += 1
                        else:
                            if heading.has_quantity:
                                worksheet.write(row, index + 1, '0', row_format)
                                worksheet.write(row, index + 2, '0.00', row_format)
                                index += 2
                            else:
                                worksheet.write(row, index + 1, '0.00', row_format)
                                index += 1
                    index += 1
                    worksheet.write(row, index, salary_detail.get_income_format(), row_format)
                    for heading in headings.filter(type='descuentos').order_by('order'):
                        salary_headings = salary_detail.salaryheadings_set.filter(headings_id=heading.id).first()
                        if salary_headings:
                            if heading.has_quantity:
                                worksheet.write(row, index + 1, salary_headings.get_cant(), row_format)
                                worksheet.write(row, index + 2, salary_headings.get_valor_format(), row_format)
                                index += 2
                            else:
                                worksheet.write(row, index + 1, salary_headings.get_valor_format(), row_format)
                                index += 1
                        else:
                            if heading.has_quantity:
                                worksheet.write(row, index + 1, '0', row_format)
                                worksheet.write(row, index + 2, '0.00', row_format)
                                index += 2
                            else:
                                worksheet.write(row, index + 1, '0.00', row_format)
                                index += 1
                    worksheet.write(row, index + 1, salary_detail.get_expenses_format(), row_format)
                    worksheet.write(row, index + 2, salary_detail.get_total_amount_format(), row_format)
                    row += 1
                workbook.close()
                output.seek(0)
                response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f"attachment; filename='PLANILLA_{datetime.now().date().strftime('%d_%m_%Y')}.xlsx'"
                return response
            elif action == 'export_template':
                year = request.POST['year']
                month = request.POST['month']
                headers = {
                    'Código': 15,
                    'Empleado': 35,
                    'Sección': 35,
                    'Cargo': 35,
                    'Número de documento': 35,
                    'Fecha de ingreso': 35,
                }
                headings = Headings.objects.filter()
                for i in headings.filter(type='haberes').order_by('order'):
                    if i.has_quantity:
                        key = f'Cantidad.{i.code.lower()}'
                        headers[key] = 45
                    headers[i.name] = 55
                headers['Subtotal'] = 50
                for i in headings.filter(type='descuentos').order_by('order'):
                    if i.has_quantity:
                        key = f'Cantidad.{i.code.lower()}'
                        headers[key] = 45
                    headers[i.name] = 55
                headers['Total Descuento'] = 40
                headers['Total a Cobrar'] = 40
                output = BytesIO()
                workbook = xlsxwriter.Workbook(output)
                worksheet = workbook.add_worksheet('plantilla')
                cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
                row_format = workbook.add_format({'align': 'center', 'border': 1})
                index = 0
                for name, width in headers.items():
                    worksheet.set_column(first_col=0, last_col=index, width=width)
                    worksheet.write(0, index, name, cell_format)
                    index += 1
                row = 1
                for employee in Employee.objects.filter(user__is_active=True):
                    worksheet.write(row, 0, employee.code, row_format)
                    worksheet.write(row, 1, employee.user.names, row_format)
                    worksheet.write(row, 2, employee.area.name, row_format)
                    worksheet.write(row, 3, employee.position.name, row_format)
                    worksheet.write(row, 4, employee.dni, row_format)
                    worksheet.write(row, 5, employee.hiring_date_format(), row_format)
                    index = 5
                    salary_detail = SalaryDetail.objects.filter(employee=employee, salary__year=year, salary__month=month).first()
                    for heading in headings.filter(type='haberes').order_by('order'):
                        salary_headings = SalaryHeadings.objects.filter(headings_id=heading.id, salary_detail__employee=employee, salary_detail__salary__year=year, salary_detail__salary__month=month).first()
                        if salary_headings:
                            if heading.has_quantity:
                                worksheet.write(row, index + 1, salary_headings.get_cant(), row_format)
                                worksheet.write(row, index + 2, salary_headings.get_valor_format(), row_format)
                                index += 2
                            else:
                                worksheet.write(row, index + 1, salary_headings.get_valor_format(), row_format)
                                index += 1
                        else:
                            if heading.has_quantity:
                                if heading.code == 'salario':
                                    worksheet.write(row, index + 1, employee.get_amount_of_assists(year, month), row_format)
                                    worksheet.write(row, index + 2, float(employee.remuneration), row_format)
                                else:
                                    worksheet.write(row, index + 1, '0', row_format)
                                    worksheet.write(row, index + 2, '0.00', row_format)
                                index += 2
                            else:
                                worksheet.write(row, index + 1, '0.00', row_format)
                                index += 1
                    index += 1
                    if salary_detail:
                        worksheet.write(row, index, salary_detail.get_income_format(), row_format)
                    else:
                        worksheet.write(row, index, '0.00', row_format)
                    for heading in headings.filter(type='descuentos').order_by('order'):
                        salary_headings = SalaryHeadings.objects.filter(headings_id=heading.id, salary_detail__employee=employee, salary_detail__salary__year=year, salary_detail__salary__month=month).first()
                        if salary_headings:
                            if heading.has_quantity:
                                worksheet.write(row, index + 1, salary_headings.get_cant(), row_format)
                                worksheet.write(row, index + 2, salary_headings.get_valor_format(), row_format)
                                index += 2
                            else:
                                worksheet.write(row, index + 1, salary_headings.get_valor_format(), row_format)
                                index += 1
                        else:
                            if heading.has_quantity:
                                worksheet.write(row, index + 1, '0', row_format)
                                worksheet.write(row, index + 2, '0.00', row_format)
                                index += 2
                            else:
                                worksheet.write(row, index + 1, '0.00', row_format)
                                index += 1
                    if salary_detail:
                        worksheet.write(row, index + 1, salary_detail.get_expenses_format(), row_format)
                        worksheet.write(row, index + 2, salary_detail.get_total_amount_format(), row_format)
                    else:
                        worksheet.write(row, index + 1, '0.00', row_format)
                        worksheet.write(row, index + 2, '0.00', row_format)
                    row += 1
                workbook.close()
                output.seek(0)
                response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f"attachment; filename='PLANILLA_{datetime.now().date().strftime('%d_%m_%Y')}.xlsx'"
                return response
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Salarios'
        context['create_url'] = reverse_lazy('salary_create')
        return context


class SalaryCreateView(GroupPermissionMixin, CreateView):
    model = Salary
    template_name = 'salary/admin/create.html'
    form_class = SalaryForm
    success_url = reverse_lazy('salary_list')
    permission_required = 'add_salary'

    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'add':
                with transaction.atomic():
                    salary = Salary.objects.get_or_create(year=int(request.POST['year']), month=int(request.POST['month']))[0]
                    for i in json.loads(request.POST['headings']):
                        heading = i
                        employee = Employee.objects.get(pk=int(heading['employee']['id']))
                        salary_detail = salary.salarydetail_set.filter(employee=employee).first()
                        if salary_detail:
                            salary_detail.salaryheadings_set.all().delete()
                        else:
                            salary_detail = SalaryDetail()
                            salary_detail.salary_id = salary.id
                            salary_detail.employee_id = employee.id
                            salary_detail.save()
                        del heading['employee']
                        del heading['total_discounts']
                        del heading['total_charge']
                        del heading['total_assets']
                        for key, value in heading.items():
                            detail = SalaryHeadings()
                            detail.salary_detail_id = salary_detail.id
                            detail.headings_id = int(value['id'])
                            detail.cant = int(value['cant'])
                            detail.valor = float(value['amount'])
                            detail.save()
                        salary_detail.income = salary_detail.salaryheadings_set.filter(headings__type='haberes').aggregate(result=Coalesce(Sum('valor'), 0.00, output_field=FloatField()))['result']
                        salary_detail.expenses = salary_detail.salaryheadings_set.filter(headings__type='descuentos').aggregate(result=Coalesce(Sum('valor'), 0.00, output_field=FloatField()))['result']
                        salary_detail.total_amount = float(salary_detail.income) - float(salary_detail.expenses)
                        salary_detail.save()
            elif action == 'search_employee':
                data = []
                term = request.POST['term']
                for i in Employee.objects.filter(Q(user__names__icontains=term) | Q(dni__icontains=term) | Q(code__icontains=term)).order_by('user__names')[0:10]:
                    item = i.toJSON()
                    item['text'] = i.get_full_name()
                    data.append(item)
            elif action == 'search_employees':
                detail = []
                year = int(request.POST['year'])
                month = int(request.POST['month'])
                employees_ids = json.loads(request.POST['employees_ids'])
                employees = Employee.objects.filter(user__is_active=True)
                if len(employees_ids):
                    employees = employees.filter(id__in=employees_ids)
                columns = [{'data': 'employee.user.names'}]
                headings = Headings.objects.filter(state=True)
                for i in headings.filter(type='haberes').order_by('type', 'order', 'has_quantity'):
                    if i.has_quantity:
                        columns.append({"data": f"{i.code}.cant"})
                    columns.append({"data": i.code})
                columns.append({"data": "total_assets"})
                for i in headings.filter(type='descuentos').order_by('type', 'order'):
                    if i.has_quantity:
                        columns.append({"data": f"{i.code}.cant"})
                    columns.append({"data": i.code})
                columns.append({"data": "total_discounts"})
                columns.append({"data": "total_charge"})
                for employee in employees:
                    heading = {}
                    for d in headings.filter(type='haberes').order_by('order'):
                        item = d.toJSON()
                        item['cant'] = 0
                        item['amount'] = 0.00
                        if d.code == 'salario':
                            item['amount'] = float(employee.remuneration)
                            item['cant'] = employee.get_amount_of_assists(year, month)
                        queryset = d.get_amount_detail_salary(employee=employee.id, year=year, month=month)
                        if queryset is not None:
                            item['amount'] = float(queryset.valor)
                            item['cant'] = queryset.cant
                        heading[d.code] = item
                    for d in headings.filter(type='descuentos').order_by('order'):
                        item = d.toJSON()
                        item['cant'] = 0
                        item['amount'] = 0.00
                        queryset = d.get_amount_detail_salary(employee=employee.id, year=year, month=month)
                        if queryset is not None:
                            item['amount'] = float(queryset.valor)
                            item['cant'] = queryset.cant
                        heading[d.code] = item
                    salary_detail = SalaryDetail.objects.filter(employee_id=employee.id, salary__year=year, salary__month=month).first()
                    if salary_detail:
                        heading['total_assets'] = {'code': 'total_assets', 'amount': float(salary_detail.income)}
                        heading['total_discounts'] = {'code': 'total_discounts', 'amount': float(salary_detail.expenses)}
                        heading['total_charge'] = {'code': 'total_charge', 'amount': float(salary_detail.total_amount)}
                    else:
                        heading['total_assets'] = {'code': 'total_assets', 'amount': 0.00}
                        heading['total_discounts'] = {'code': 'total_discounts', 'amount': 0.00}
                        heading['total_charge'] = {'code': 'total_charge', 'amount': float(employee.remuneration)}
                    heading['employee'] = employee.toJSON()
                    detail.append(heading)
                data = {'detail': detail, 'columns': columns}
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['title'] = 'Nuevo registro de un Salario'
        context['list_url'] = self.success_url
        context['action'] = 'add'
        context['assets'] = Headings.objects.filter(state=True, type='haberes').order_by('id')
        context['discounts'] = Headings.objects.filter(state=True, type='descuentos').order_by('id')
        return context


class SalaryDeleteView(GroupPermissionMixin, TemplateView):
    template_name = 'salary/admin/delete.html'
    success_url = reverse_lazy('salary_list')
    permission_required = 'delete_salary'

    def get(self, request, *args, **kwargs):
        if self.get_object() is not None:
            return super(SalaryDeleteView, self).get(request, *args, **kwargs)
        messages.error(request, 'No existen salarios en ese año y mes')
        return HttpResponseRedirect(self.success_url)

    def get_object(self):
        queryset = None
        try:
            month = self.request.session['salaries']['month']
            employees = self.request.session['salaries']['employees']
            queryset = SalaryDetail.objects.filter(salary__year=self.request.session['salaries']['year'])
            if len(month):
                queryset = queryset.filter(salary__month=month['id'])
            if len(employees):
                queryset = queryset.filter(employee_id__in=employees)
        except:
            pass
        return queryset

    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'remove':
                self.get_object().delete()
                if 'salaries' in request.session:
                    del request.session['salaries']
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Notificación de eliminación'
        context['list_url'] = self.success_url
        return context


class SalaryPrintReceiptView(LoginRequiredMixin, View):
    success_url = reverse_lazy('salary_list')

    def get_success_url(self):
        if self.request.user.is_employee():
            return reverse_lazy('salary_employee_list')
        return self.success_url

    def get(self, request, *args, **kwargs):
        try:
            salary_detail = SalaryDetail.objects.get(pk=self.kwargs['pk'])
            context = {
                'salary_detail': salary_detail,
                'company': request.tenant.company,
                'prints': [1, 2],
                'date_joined': datetime.now().date()
            }
            pdf_file = printer.create_pdf(context=context, template_name='salary/format/format1.html')
            return HttpResponse(pdf_file, content_type='application/pdf')
        except:
            pass
        return HttpResponseRedirect(self.get_success_url())


class SalaryEmployeeListView(GroupPermissionMixin, FormView):
    form_class = SalaryForm
    template_name = 'salary/employee/list.html'
    permission_required = 'view_employee_salary'

    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'search':
                data = []
                year = request.POST['year']
                month = request.POST['month']
                queryset = SalaryDetail.objects.filter(employee=request.user.employee)
                if len(year):
                    queryset = queryset.filter(salary__year=year)
                if len(month):
                    queryset = queryset.filter(salary__month=month)
                for i in queryset:
                    data.append(i.toJSON())
            elif action == 'search_detail_headings':
                data = []
                salary_detail = SalaryDetail.objects.get(pk=request.POST['id'])
                for i in salary_detail.salaryheadings_set.filter(headings__type='haberes', valor__gt=0).order_by('headings__order'):
                    data.append([i.headings.name, i.get_cant(), i.get_valor_format(), '---'])
                for i in salary_detail.salaryheadings_set.filter(headings__type='descuentos', valor__gt=0).order_by('headings__order'):
                    data.append([i.headings.name, i.get_cant(), '---', i.get_valor_format()])
                data.append(['Subtotal de Ingresos', '---', '---', salary_detail.get_income_format()])
                data.append(['Subtotal de Descuentos', '---', '---', salary_detail.get_expenses_format()])
                data.append(['Total a recibir', '---', '---', salary_detail.get_total_amount_format()])
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Listado de Salarios'
        return context
